
import re
import unicodedata
from io import BytesIO
from datetime import date

import pandas as pd
import numpy as np
import streamlit as st

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# Helpers gerais
# =========================================================

def normalize_colname(name: str) -> str:
    if not isinstance(name, str):
        return ""
    s = name.strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def norm_person_key(name: str) -> str:
    """Chave robusta p/ casar nomes entre relatórios."""
    if name is None:
        return ""
    s = str(name).strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[^A-Z ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def display_name_from_full(name: str) -> str:
    """
    Replica o “jeito” do print:
    - se o nome tiver até 3 palavras: usa todas
    - se tiver mais: usa só as 2 primeiras
    """
    if not name:
        return ""
    s = str(name).strip()
    s = re.sub(r"\s+", " ", s)
    parts = s.split(" ")
    if len(parts) <= 3:
        out = " ".join(parts)
    else:
        out = " ".join(parts[:2])

    # Title case “esperto” (mantém acentos se existirem no original)
    return out.title()


def parse_time_to_timedelta(value):
    if pd.isna(value):
        return pd.NaT

    s = str(value).strip()
    if s == "" or s in {"-", "—", "–", "--", "nan", "None"}:
        return pd.NaT

    if isinstance(value, pd.Timedelta):
        return value

    if re.match(r"^\d{1,2}:\d{2}$", s):
        try:
            h, m = map(int, s.split(":"))
            return pd.Timedelta(hours=h, minutes=m)
        except Exception:
            return pd.NaT

    if re.match(r"^\d{1,2}:\d{2}:\d{2}$", s):
        try:
            h, m, sec = map(int, s.split(":"))
            return pd.Timedelta(hours=h, minutes=m, seconds=sec)
        except Exception:
            return pd.NaT

    td = pd.to_timedelta(s, errors="coerce")
    if not pd.isna(td):
        return td

    try:
        f = float(s.replace(",", "."))
        return pd.to_timedelta(f, unit="D")
    except Exception:
        return pd.NaT


def fmt_td_hms(td: pd.Timedelta) -> str:
    """Formato do print: H:MM:SS (sem zero à esquerda na hora)."""
    if pd.isna(td):
        return "0:00:00"
    sec = int(td.total_seconds())
    if sec < 0:
        sec = 0
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return f"{h}:{m:02d}:{s:02d}"


def fmt_td_hms_signed(td: pd.Timedelta) -> str:
    if pd.isna(td):
        return "0:00:00"
    sec = int(td.total_seconds())
    sign = "-" if sec < 0 else ""
    sec = abs(sec)
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return f"{sign}{h}:{m:02d}:{s:02d}"


def fmt_dt_time(dt) -> str:
    """INÍCIO/FIM no formato HH:MM:SS (ou vazio)."""
    if pd.isna(dt) or dt is None:
        return ""
    try:
        ts = pd.Timestamp(dt)
        return ts.strftime("%H:%M:%S")
    except Exception:
        return ""


# =========================================================
# Mapeamento de colunas (Horas Extras)
# =========================================================

CANONICAL_MAP = {
    "colaborador": "Colaborador",
    "cpf": "CPF",
    "data": "Data",
    "previstas": "Previstas",
    "trabalhadas": "Trabalhadas",
    "abonadas": "Abonadas",
    "atrasos": "Atrasos",
    "noturnas": "Noturnas",
    "hr. ficta": "Horas Fictas",
    "hr ficta": "Horas Fictas",
    "horas fictas": "Horas Fictas",
    "faltas": "Faltas",
}


# =========================================================
# Etapa 1 — Detectar cabeçalhos e montar tabela longa (Horas Extras)
# =========================================================

def _build_long_from_folha_ponto(raw: pd.DataFrame) -> pd.DataFrame:
    mask_blocks = raw.apply(
        lambda r: r.astype(str).str.contains("DADOS DO COLABORADOR", na=False).any(),
        axis=1,
    )
    block_starts = [i for i in raw.index if mask_blocks[i]]
    if not block_starts:
        return pd.DataFrame()

    n_rows, _ = raw.shape
    rows = []

    def find_header_row(start, end):
        for i in range(start, end):
            if raw.iloc[i].astype(str).str.contains("DIA / MÊS", na=False).any():
                return i
        return None

    def find_col_idx(header_row, label):
        row = raw.iloc[header_row].astype(str)
        hits = np.where(row.str.contains(label, case=False, na=False))[0]
        return int(hits[0]) if hits.size else None

    def extract_value_after_label(block_start, block_end, label):
        for i in range(block_start, block_end):
            row = raw.iloc[i]
            srow = row.astype(str)
            for j, val in enumerate(srow):
                if label in str(val):
                    for k in range(j + 1, len(row)):
                        v = row[k]
                        if (not pd.isna(v)) and str(v).strip() != "":
                            return v
        return None

    for idx, start in enumerate(block_starts):
        end = block_starts[idx + 1] if idx + 1 < len(block_starts) else n_rows

        header_row = find_header_row(start, end)
        if header_row is None:
            continue

        col_data = find_col_idx(header_row, "DIA / MÊS")
        col_pontos = find_col_idx(header_row, "PONTOS")
        col_trab = find_col_idx(header_row, "TRABALHADAS")
        col_abono = find_col_idx(header_row, "ABONO")
        col_prev = find_col_idx(header_row, "PREVISTAS")

        nome_val = extract_value_after_label(start, header_row, "Nome:")
        cpf_val = extract_value_after_label(start, header_row, "CPF:")

        cpf_str = ""
        if cpf_val is not None:
            try:
                if isinstance(cpf_val, (int, float)):
                    cpf_str = str(int(cpf_val))
                else:
                    cpf_str = str(cpf_val).strip()
            except Exception:
                cpf_str = str(cpf_val).strip()

        r = header_row + 1
        while r < end:
            v_data = raw.iloc[r, col_data] if col_data is not None else None
            if isinstance(v_data, str) and v_data.strip().startswith("Total"):
                break
            if pd.isna(v_data):
                r += 1
                continue

            data_val = v_data
            pontos_val = raw.iloc[r, col_pontos] if col_pontos is not None else None
            trab_val = raw.iloc[r, col_trab] if col_trab is not None else None
            abono_val = raw.iloc[r, col_abono] if col_abono is not None else None
            prev_val = raw.iloc[r, col_prev] if col_prev is not None else None

            faltas_val = None
            if isinstance(pontos_val, str) and "FALTA" in pontos_val.upper():
                faltas_val = prev_val

            rows.append(
                {
                    "Colaborador": str(nome_val).strip() if nome_val is not None else "",
                    "CPF": cpf_str,
                    "Data": data_val,
                    "Previstas": prev_val,
                    "Trabalhadas": trab_val,
                    "Abonadas": abono_val,
                    "Atrasos": None,
                    "Noturnas": None,
                    "Horas Fictas": None,
                    "Faltas": faltas_val,
                }
            )
            r += 1

    return pd.DataFrame(rows)


def _build_long_from_solides(raw: pd.DataFrame) -> pd.DataFrame:
    mask_previstas = raw.apply(
        lambda row: row.astype(str).str.contains("Previstas", case=False, na=False).any(),
        axis=1,
    )
    header_idx = [i for i in raw.index if mask_previstas[i]]
    if not header_idx:
        return pd.DataFrame()

    blocks = []
    for pos, h in enumerate(header_idx):
        next_h = header_idx[pos + 1] if pos + 1 < len(header_idx) else len(raw)
        header_row = raw.iloc[h]
        block = raw.iloc[h + 1: next_h].copy()
        block = block.dropna(how="all")
        if block.empty:
            continue
        block.columns = header_row.values
        blocks.append(block)

    if not blocks:
        return pd.DataFrame()

    return pd.concat(blocks, ignore_index=True)


def detect_blocks_and_build_long(raw: pd.DataFrame) -> pd.DataFrame:
    long_df = _build_long_from_folha_ponto(raw)
    if not long_df.empty:
        return long_df

    long_df = _build_long_from_solides(raw)
    if not long_df.empty:
        return long_df

    raise ValueError(
        "Não consegui identificar o layout do relatório.\n"
        "- Para Folha de Ponto, preciso da área 'DADOS DO COLABORADOR' com o quadro DIA / MÊS.\n"
        "- Para o modelo antigo, preciso das colunas com cabeçalho 'Previstas', 'Trabalhadas' etc."
    )


# =========================================================
# Etapa 2 — Semanas 26→25
# =========================================================

def classify_weeks(dates: pd.Series) -> pd.Series:
    if dates.empty:
        return pd.Series(dtype="object")

    dts = pd.to_datetime(dates, errors="coerce")
    dts_norm = dts.dt.normalize()

    ym_list = sorted({(d.year, d.month) for d in dts_norm.dropna()})
    if not ym_list:
        return pd.Series(["Semana 1"] * len(dts), index=dates.index)

    prev_year, prev_month = ym_list[0]
    if len(ym_list) > 1:
        curr_year, curr_month = ym_list[1]
    else:
        curr_year, curr_month = prev_year, prev_month

    def _classify(d):
        if pd.isna(d):
            return None

        y, m, day = d.year, d.month, d.day

        if (y == prev_year) and (m == prev_month):
            return "Semana 1"

        if (y == curr_year) and (m == curr_month):
            if 1 <= day <= 7:
                return "Semana 2"
            elif 8 <= day <= 14:
                return "Semana 3"
            elif 15 <= day <= 21:
                return "Semana 4"
            else:
                return "Semana 5"

        return "Semana 5"

    return dts_norm.apply(_classify)


# =========================================================
# Etapa 2.1 — Feriados nacionais e domingos (HE 100%)
# =========================================================

NATIONAL_HOLIDAYS = {
    (1, 1),
    (4, 21),
    (5, 1),
    (9, 7),
    (10, 12),
    (11, 2),
    (11, 15),
    (11, 20),
    (12, 25),
}


def is_national_holiday(date_value) -> bool:
    if pd.isna(date_value):
        return False
    d = pd.Timestamp(date_value)
    return (d.month, d.day) in NATIONAL_HOLIDAYS


# =========================================================
# Etapa 3 — Limpeza e cálculo HE (Horas Extras)
# =========================================================

def clean_and_enrich(long_df: pd.DataFrame) -> pd.DataFrame:
    df = long_df.dropna(axis=1, how="all").copy()

    rename_map = {}
    for col in df.columns:
        canon = CANONICAL_MAP.get(normalize_colname(col))
        if canon:
            rename_map[col] = canon
    df = df.rename(columns=rename_map)

    required = ["Colaborador", "CPF", "Data", "Previstas", "Trabalhadas"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError("Colunas obrigatórias ausentes: " + ", ".join(missing))

    for opt in ["Abonadas", "Atrasos", "Noturnas", "Horas Fictas", "Faltas"]:
        if opt not in df.columns:
            df[opt] = pd.NA

    df = df[
        [
            "Colaborador",
            "CPF",
            "Data",
            "Previstas",
            "Trabalhadas",
            "Abonadas",
            "Atrasos",
            "Noturnas",
            "Horas Fictas",
            "Faltas",
        ]
    ].copy()

    df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
    df = df[~df["Data"].isna()].copy()

    df["Colaborador"] = df["Colaborador"].astype(str).str.strip()
    df["CPF"] = df["CPF"].astype(str).str.strip()

    time_cols = [
        "Previstas",
        "Trabalhadas",
        "Abonadas",
        "Atrasos",
        "Noturnas",
        "Horas Fictas",
        "Faltas",
    ]
    for col in time_cols:
        df[f"{col}_td"] = df[col].apply(parse_time_to_timedelta)
        df[f"{col}_td"] = df[f"{col}_td"].fillna(pd.Timedelta(0))

    zero = pd.Timedelta(0)

    # Atrasos: Trabalhadas - Previstas (se negativo)
    df["Atrasos_td"] = df["Trabalhadas_td"] - df["Previstas_td"]
    df["Atrasos_td"] = df["Atrasos_td"].where(df["Atrasos_td"] < zero, zero)

    # Excedente: positivo
    df["Excedente_td"] = df["Trabalhadas_td"] - df["Previstas_td"]
    df["Excedente_td"] = df["Excedente_td"].where(df["Excedente_td"] > zero, zero)

    df["is_domingo"] = df["Data"].dt.dayofweek == 6
    df["is_feriado"] = df["Data"].apply(is_national_holiday)
    mask_especial = df["is_domingo"] | df["is_feriado"]
    mask_normal = ~mask_especial

    df["HE50_td"] = zero
    df["HE70_td"] = zero
    df["HE100_td"] = zero

    two_hours = pd.to_timedelta(2, unit="h")

    # Normal: 50% até 2h, resto 70%
    df.loc[mask_normal, "HE50_td"] = df.loc[mask_normal, "Excedente_td"].clip(
        lower=zero, upper=two_hours
    )
    df.loc[mask_normal, "HE70_td"] = (
        df.loc[mask_normal, "Excedente_td"] - df.loc[mask_normal, "HE50_td"]
    ).clip(lower=zero)

    # Domingo/feriado: 100%
    df.loc[mask_especial, ["HE50_td", "HE70_td"]] = zero
    df.loc[mask_especial, "HE100_td"] = df.loc[mask_especial, "Excedente_td"]

    # Sem previstas e trabalhou (escala 0): 100%
    mask_sem_previstas = (
        (df["Previstas_td"] == zero)
        & (df["Trabalhadas_td"] > zero)
        & (~mask_especial)
    )
    df.loc[mask_sem_previstas, ["HE50_td", "HE70_td"]] = zero
    df.loc[mask_sem_previstas, "HE100_td"] = df.loc[mask_sem_previstas, "Excedente_td"]

    # Abono sem trabalho: zera tudo
    mask_abono_sem_trabalho = (df["Abonadas_td"] > zero) & (df["Trabalhadas_td"] == zero)
    df.loc[
        mask_abono_sem_trabalho,
        ["Atrasos_td", "Excedente_td", "HE50_td", "HE70_td", "HE100_td"],
    ] = zero

    # Abono com trabalho: tudo trabalhado vira 100%
    mask_abono_com_trabalho = (df["Abonadas_td"] > zero) & (df["Trabalhadas_td"] > zero)
    df.loc[mask_abono_com_trabalho, "Atrasos_td"] = zero
    df.loc[mask_abono_com_trabalho, "Excedente_td"] = df.loc[
        mask_abono_com_trabalho, "Trabalhadas_td"
    ]
    df.loc[mask_abono_com_trabalho, ["HE50_td", "HE70_td"]] = zero
    df.loc[mask_abono_com_trabalho, "HE100_td"] = df.loc[
        mask_abono_com_trabalho, "Excedente_td"
    ]

    df["Semana"] = classify_weeks(df["Data"])
    df = df.sort_values(["Colaborador", "Data"]).reset_index(drop=True)

    # chave p/ cruzar com o outro relatório
    df["ColabKey"] = df["Colaborador"].apply(norm_person_key)
    return df


# =========================================================
# Abas Semanais e Resumo (mantidas)
# =========================================================

def build_weekly_sheets(df: pd.DataFrame) -> dict:
    weeks = {}
    zero = pd.Timedelta(0)

    def format_timedelta(td: pd.Timedelta) -> str:
        if pd.isna(td):
            return "00:00"
        total_seconds = int(td.total_seconds())
        if total_seconds < 0:
            total_seconds = 0
        hours, remainder = divmod(total_seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}"

    def format_timedelta_signed(td: pd.Timedelta) -> str:
        if pd.isna(td):
            return "00:00"
        total_seconds = int(td.total_seconds())
        sign = "-" if total_seconds < 0 else ""
        total_seconds = abs(total_seconds)
        hours, remainder = divmod(total_seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        return f"{sign}{hours:02d}:{minutes:02d}"

    for i in range(1, 6):
        wname = f"Semana {i}"
        sub = df[df["Semana"] == wname].copy()

        if sub.empty:
            out = pd.DataFrame(
                columns=[
                    "Semana", "Colaborador", "CPF", "Data",
                    "Previstas", "Trabalhadas", "Atrasos",
                    "HE 50%", "HE 70%", "HE 100%", "Noturnas",
                ]
            )
            weeks[wname] = out
            continue

        sub = sub.sort_values(["Colaborador", "CPF", "Data"]).reset_index(drop=True)

        linhas = []
        for (colab, cpf), g in sub.groupby(["Colaborador", "CPF"], sort=False):
            for _, r in g.iterrows():
                worked_display = r["Trabalhadas_td"]
                if (r["Abonadas_td"] > zero) and (r["Trabalhadas_td"] == zero):
                    worked_display = r["Abonadas_td"]

                linhas.append(
                    {
                        "Semana": r["Semana"],
                        "Colaborador": r["Colaborador"],
                        "CPF": r["CPF"],
                        "Data": r["Data"].date(),
                        "Previstas": format_timedelta(r["Previstas_td"]),
                        "Trabalhadas": format_timedelta(worked_display),
                        "Atrasos": format_timedelta_signed(r["Atrasos_td"]),
                        "HE 50%": format_timedelta(r["HE50_td"]),
                        "HE 70%": format_timedelta(r["HE70_td"]),
                        "HE 100%": format_timedelta(r["HE100_td"]),
                        "Noturnas": format_timedelta(r["Noturnas_td"]),
                    }
                )

            soma_prev = g["Previstas_td"].sum()
            soma_atra = g["Atrasos_td"].sum()
            soma_he50 = g["HE50_td"].sum()
            soma_he70 = g["HE70_td"].sum()
            soma_he100 = g["HE100_td"].sum()
            soma_not = g["Noturnas_td"].sum()

            worked_display_series = g["Trabalhadas_td"].copy()
            mask_abono_sem_trab = (g["Abonadas_td"] > zero) & (g["Trabalhadas_td"] == zero)
            worked_display_series = worked_display_series.where(~mask_abono_sem_trab, g["Abonadas_td"])
            soma_trab_display = worked_display_series.sum()

            total50_semana_td = soma_he50 + soma_atra  # 50 “líquido” (com atraso)

            linhas.append(
                {
                    "Semana": wname,
                    "Colaborador": f"{colab} - TOTAL",
                    "CPF": cpf,
                    "Data": "",
                    "Previstas": format_timedelta(soma_prev),
                    "Trabalhadas": format_timedelta(soma_trab_display),
                    "Atrasos": format_timedelta_signed(soma_atra),
                    "HE 50%": format_timedelta_signed(total50_semana_td),
                    "HE 70%": format_timedelta(soma_he70),
                    "HE 100%": format_timedelta(soma_he100),
                    "Noturnas": format_timedelta(soma_not),
                }
            )

        weeks[wname] = pd.DataFrame(linhas)

    return weeks


def build_resumo(df: pd.DataFrame) -> pd.DataFrame:
    grp = df.groupby(["Colaborador", "CPF"], as_index=False)[
        [
            "Previstas_td",
            "Trabalhadas_td",
            "Abonadas_td",
            "HE50_td",
            "HE70_td",
            "HE100_td",
            "Atrasos_td",
            "Faltas_td",
            "Noturnas_td",
        ]
    ].sum()

    def format_timedelta(td: pd.Timedelta) -> str:
        if pd.isna(td):
            return "00:00"
        total_seconds = int(td.total_seconds())
        if total_seconds < 0:
            total_seconds = 0
        hours, remainder = divmod(total_seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        return f"{hours:02d}:{minutes:02d}"

    def format_timedelta_signed(td: pd.Timedelta) -> str:
        if pd.isna(td):
            return "00:00"
        total_seconds = int(td.total_seconds())
        sign = "-" if total_seconds < 0 else ""
        total_seconds = abs(total_seconds)
        hours, remainder = divmod(total_seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        return f"{sign}{hours:02d}:{minutes:02d}"

    atrasos_total_td = grp["Atrasos_td"]
    grp["Total50_td"] = grp["HE50_td"] + atrasos_total_td
    grp["Total70_td"] = grp["HE70_td"]
    grp["Total100_td"] = grp["HE100_td"]
    grp["Geral_td"] = grp["Total50_td"] + grp["Total70_td"] + grp["Total100_td"]
    grp["Meio_Geral_td"] = grp["Geral_td"] / 2
    grp["Setenta_Cem_td"] = grp["Total70_td"] + grp["Total100_td"]
    grp["Cinq_Pagar_td"] = grp["Meio_Geral_td"] - grp["Setenta_Cem_td"]
    trab_com_abono_td = grp["Trabalhadas_td"] + grp["Abonadas_td"]

    out = pd.DataFrame()
    out["Colaborador"] = grp["Colaborador"]
    out["CPF"] = grp["CPF"]
    out["Previstas(D)"] = grp["Previstas_td"].apply(format_timedelta)
    out["Trabalhadas (E)"] = trab_com_abono_td.apply(format_timedelta)
    out["Atrasos (F)"] = atrasos_total_td.apply(format_timedelta_signed)
    out["Total 50 (G)"] = grp["Total50_td"].apply(format_timedelta_signed)
    out["Total 70 (H)"] = grp["Total70_td"].apply(format_timedelta)
    out["Total 100 (I)"] = grp["Total100_td"].apply(format_timedelta)
    out["GERAL (J)"] = grp["Geral_td"].apply(format_timedelta)
    out["Noturnas (K)"] = grp["Noturnas_td"].apply(format_timedelta)
    out["1/2 GERAL"] = grp["Meio_Geral_td"].apply(format_timedelta)
    out["70 + 100"] = grp["Setenta_Cem_td"].apply(format_timedelta)
    out["50 À PG"] = grp["Cinq_Pagar_td"].apply(format_timedelta_signed)

    return out.sort_values("Colaborador").reset_index(drop=True)


# =========================================================
# Parser do relatório “Pontos com Endereço”
# =========================================================

def read_pontos_endereco(uploaded_file) -> pd.DataFrame:
    """
    Lê o relatório bruto “Relatório de pontos com endereço” e devolve
    um DF diário com:
    ColabKey, Colaborador (display), Data, Inicio_dt, Fim_dt, Base
    """
    raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)

    header_row = None
    for i in range(min(40, len(raw))):
        row = raw.iloc[i].astype(str)
        if row.str.contains(r"\bColaborador\b", case=False, na=False).any():
            header_row = i
            break
    if header_row is None:
        raise ValueError("Não encontrei o cabeçalho do relatório de pontos com endereço (coluna 'Colaborador').")

    df = pd.read_excel(uploaded_file, sheet_name=0, header=header_row)
    df = df.dropna(axis=1, how="all").copy()
    df.columns = [str(c).strip() for c in df.columns]

    col_colab = [c for c in df.columns if "Colaborador" in c][0]
    data_cols = [c for c in df.columns if str(c).strip().startswith("Data")]
    if not data_cols:
        raise ValueError("Não encontrei colunas 'Data' no relatório de pontos com endereço.")
    col_data_inicio = data_cols[0]
    col_data_fim = None
    for c in data_cols[1:]:
        if str(c).startswith("Data.") or str(c).strip() != "Data":
            col_data_fim = c
            break
    if col_data_fim is None and len(data_cols) >= 2:
        col_data_fim = data_cols[1]

    endereco_cols = [c for c in df.columns if "Endereço" in c or "Endereco" in c]
    if not endereco_cols:
        raise ValueError("Não encontrei coluna(s) de Endereço no relatório de pontos com endereço.")

    # normaliza nomes
    df[col_colab] = (
        df[col_colab]
        .astype(str)
        .replace("nan", "", regex=False)
        .str.replace(r"[\n\r\t]+", " ", regex=True)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    # datas
    df[col_data_inicio] = pd.to_datetime(df[col_data_inicio], errors="coerce")
    if col_data_fim:
        df[col_data_fim] = pd.to_datetime(df[col_data_fim], errors="coerce")
    else:
        df[col_data_fim] = pd.NaT

    # mantém linhas com alguma marcação
    df = df[df[col_data_inicio].notna() | df[col_data_fim].notna()].copy()

    # se não houver início, usa fim como início (não perde o dia)
    df[col_data_inicio] = df[col_data_inicio].fillna(df[col_data_fim])

    # base por linha: primeiro endereço não vazio (entre as colunas de endereço)
    end_tmp = df[endereco_cols].copy()
    end_tmp = end_tmp.replace({np.nan: None})
    df["_end_row"] = end_tmp.bfill(axis=1).iloc[:, 0].fillna("")

    df["_data"] = df[col_data_inicio].dt.date
    df["_key"] = df[col_colab].apply(norm_person_key)
    df["_disp"] = df[col_colab].apply(display_name_from_full)

    # Ordena por início p/ pegar base do primeiro ponto do dia
    df = df.sort_values(by=[col_colab, col_data_inicio])

    # Agrega por colaborador e data
    daily = (
        df.groupby(["_key", "_disp", "_data"], as_index=False)
        .agg(
            Inicio_dt=(col_data_inicio, "min"),
            Fim_dt=(col_data_fim, "max"),
            Base=("_end_row", "first"),
        )
    )

    daily = daily.rename(columns={"_key": "ColabKey", "_disp": "Colaborador", "_data": "Data"})

    # Se Fim_dt estiver todo NaT em algum dia, vira vazio na exibição depois
    return daily


# =========================================================
# Matinal (formato do print)
# =========================================================

def build_he_daily(df_he: pd.DataFrame) -> pd.DataFrame:
    """HE por dia p/ cruzar com o Matinal."""
    daily = (
        df_he.groupby(["ColabKey", "Data"], as_index=False)[
            ["HE50_td", "HE70_td", "HE100_td", "Atrasos_td"]
        ].sum()
    )
    daily["HE50_liq_td"] = daily["HE50_td"] + daily["Atrasos_td"]  # 50% “líquido” com sinal
    daily["HE70_100_td"] = daily["HE70_td"] + daily["HE100_td"]
    return daily[["ColabKey", "Data", "HE50_liq_td", "HE70_100_td"]]


def build_matinal_rows(df_addr_daily: pd.DataFrame, df_he_daily: pd.DataFrame, last_n_days: int = 2):
    dates = sorted(df_addr_daily["Data"].dropna().unique().tolist())
    if not dates:
        return [], [], None

    if last_n_days and len(dates) > last_n_days:
        dates = dates[-last_n_days:]

    last_date = dates[-1]

    addr = df_addr_daily[df_addr_daily["Data"].isin(dates)].copy()
    he = df_he_daily[df_he_daily["Data"].isin(dates)].copy()

    # merge por dia
    merged = addr.merge(he, how="left", on=["ColabKey", "Data"])

    # garante timedeltas
    merged["HE50_liq_td"] = merged["HE50_liq_td"].fillna(pd.Timedelta(0))
    merged["HE70_100_td"] = merged["HE70_100_td"].fillna(pd.Timedelta(0))

    # acumulado
    acc = (
        merged.groupby(["ColabKey", "Colaborador"], as_index=False)[["HE50_liq_td", "HE70_100_td"]]
        .sum()
        .rename(columns={"HE50_liq_td": "ACC50", "HE70_100_td": "ACC70100"})
    )

    # index rápido por (key, data)
    idx = {(r.ColabKey, r.Data): r for r in merged.itertuples(index=False)}

    rows = []
    for r in acc.itertuples(index=False):
        key = r.ColabKey
        disp = r.Colaborador

        row = {"COLABORADOR": disp}

        # dias completos (todos menos o último): inicio, fim, base, 50, 70+100
        for d in dates[:-1]:
            rec = idx.get((key, d))
            row[f"{d}_INICIO"] = fmt_dt_time(rec.Inicio_dt) if rec else ""
            row[f"{d}_FIM"] = fmt_dt_time(rec.Fim_dt) if rec else ""
            row[f"{d}_BASE"] = (rec.Base if rec else "") or ""
            row[f"{d}_50"] = fmt_td_hms_signed(rec.HE50_liq_td) if rec else "0:00:00"
            row[f"{d}_70100"] = fmt_td_hms(rec.HE70_100_td) if rec else "0:00:00"

        # último dia (matinal): inicio e base
        d = last_date
        rec = idx.get((key, d))
        row[f"{d}_INICIO"] = fmt_dt_time(rec.Inicio_dt) if rec else ""
        row[f"{d}_BASE"] = (rec.Base if rec else "") or ""

        # acumulado
        row["ACUM_50"] = fmt_td_hms_signed(r.ACC50)
        row["ACUM_70100"] = fmt_td_hms(r.ACC70100)

        rows.append(row)

    # ordena por nome
    rows = sorted(rows, key=lambda x: x.get("COLABORADOR", ""))

    return rows, dates, last_date


def write_matinal_sheet(workbook, sheet_name: str, rows: list, dates: list, last_date):
    # estilos
    purple = PatternFill("solid", fgColor="4C208E")
    white_bold = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=18)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = workbook.create_sheet(sheet_name)

    # total de colunas conforme layout:
    # 1 (Nome) + 5*(n-1) + 2 (último dia) + 2 (acumulado)
    n = len(dates)
    total_cols = 1 + (5 * (n - 1) if n >= 2 else 0) + 2 + 2
    last_col_letter = get_column_letter(total_cols)

    # Linha 1: título
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "Relatório Matinal"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # Linha 2: cabeçalho grupos
    r_group = 2
    r_sub = 3

    # Grupo Nome
    ws.cell(r_group, 1, "Nome")
    ws.cell(r_sub, 1, "COLABORADOR")

    # pinta cabeçalho (2 e 3)
    for rr in [r_group, r_sub]:
        c = ws.cell(rr, 1)
        c.fill = purple
        c.font = white_bold
        c.alignment = center
        c.border = border

    col = 2

    # Datas completas (todas menos a última)
    for d in dates[:-1]:
        dtxt = pd.Timestamp(d).strftime("%d/%m/%Y")
        # merge do grupo
        ws.merge_cells(start_row=r_group, start_column=col, end_row=r_group, end_column=col + 4)
        ws.cell(r_group, col, dtxt)

        subs = ["INÍCIO", "FIM", "BASE", "50%", "70% e 100%"]
        for j, s in enumerate(subs):
            ws.cell(r_sub, col + j, s)

        col += 5

    # Última data (matinal parcial)
    if last_date is not None:
        dtxt = pd.Timestamp(last_date).strftime("%d/%m/%Y")
        ws.merge_cells(start_row=r_group, start_column=col, end_row=r_group, end_column=col + 1)
        ws.cell(r_group, col, dtxt)

        ws.cell(r_sub, col, "INÍCIO")
        ws.cell(r_sub, col + 1, "BASE")
        col += 2

    # Acumulado
    ws.merge_cells(start_row=r_group, start_column=col, end_row=r_group, end_column=col + 1)
    ws.cell(r_group, col, "ACUMULADO")
    ws.cell(r_sub, col, "50%")
    ws.cell(r_sub, col + 1, "70% e 100%")

    # Estiliza cabeçalhos (linha 2 e 3)
    for rr in [r_group, r_sub]:
        for cc in range(1, total_cols + 1):
            cell = ws.cell(rr, cc)
            cell.fill = purple
            cell.font = white_bold
            cell.alignment = center
            cell.border = border

    # Ajuste de largura
    ws.column_dimensions["A"].width = 28
    for cc in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(cc)].width = 14

    # Altura cabeçalhos
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20

    # Dados
    start_row = 4
    for i, row in enumerate(rows):
        rr = start_row + i
        # Colaborador
        c = ws.cell(rr, 1, row.get("COLABORADOR", ""))
        c.alignment = left
        c.border = border

        col = 2

        # datas completas
        for d in dates[:-1]:
            ws.cell(rr, col + 0, row.get(f"{d}_INICIO", "")).alignment = center
            ws.cell(rr, col + 1, row.get(f"{d}_FIM", "")).alignment = center
            ws.cell(rr, col + 2, row.get(f"{d}_BASE", "")).alignment = center
            ws.cell(rr, col + 3, row.get(f"{d}_50", "0:00:00")).alignment = center
            ws.cell(rr, col + 4, row.get(f"{d}_70100", "0:00:00")).alignment = center
            for j in range(5):
                ws.cell(rr, col + j).border = border
            col += 5

        # última data
        if last_date is not None:
            ws.cell(rr, col + 0, row.get(f"{last_date}_INICIO", "")).alignment = center
            ws.cell(rr, col + 1, row.get(f"{last_date}_BASE", "")).alignment = center
            ws.cell(rr, col + 0).border = border
            ws.cell(rr, col + 1).border = border
            col += 2

        # acumulado
        ws.cell(rr, col + 0, row.get("ACUM_50", "0:00:00")).alignment = center
        ws.cell(rr, col + 1, row.get("ACUM_70100", "0:00:00")).alignment = center
        ws.cell(rr, col + 0).border = border
        ws.cell(rr, col + 1).border = border

    # Congela cabeçalho
    ws.freeze_panes = "B4"


# =========================================================
# Geração do Excel final
# =========================================================

def generate_excel_bytes(df_he_final: pd.DataFrame, df_addr_daily: pd.DataFrame, last_n_days_matinal: int = 2) -> bytes:
    weeks = build_weekly_sheets(df_he_final)
    resumo = build_resumo(df_he_final)

    he_daily = build_he_daily(df_he_final)
    rows, dates, last_date = build_matinal_rows(df_addr_daily, he_daily, last_n_days=last_n_days_matinal)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Semanais
        for wname, wdf in weeks.items():
            wdf.to_excel(writer, sheet_name=wname, index=False)

        # Resumo
        resumo.to_excel(writer, sheet_name="Resumo Geral", index=False)

        # Matinal (formatado)
        wb = writer.book
        write_matinal_sheet(wb, "Matinal", rows, dates, last_date)

        # remove a sheet padrão “Sheet” se existir
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    output.seek(0)
    return output.getvalue()


# =========================================================
# Interface Streamlit
# =========================================================

def main():
    st.set_page_config(page_title="WS | Processador de Jornada (HE + Matinal)", layout="wide")
    st.title("📊 Processador de Jornada — Horas Extras + Matinal (no formato do print)")

    st.markdown(
        """
**O que este app faz:**
- Lê o **relatório de Horas Extras / Folha de Ponto** e calcula **HE 50% / 70% / 100%**.
- Lê o **Relatório de pontos com endereço** e monta **INÍCIO / FIM / BASE**.
- Gera o Excel com:
  - **Semana 1..5**
  - **Resumo Geral**
  - **Matinal** (agrupado por dia, com **ACUMULADO**, igual ao print)
        """
    )

    c1, c2 = st.columns(2)
    with c1:
        horas_file = st.file_uploader("1) Envie o relatório de Horas Extras / Folha de Ponto (.xlsx)", type=["xlsx"], key="he")
    with c2:
        endereco_file = st.file_uploader("2) Envie o Relatório de Pontos com Endereço (.xlsx)", type=["xlsx"], key="end")

    if not horas_file or not endereco_file:
        st.info("⬆️ Envie os **dois arquivos** para eu montar o Matinal no formato do print.")
        return

    last_n_days_matinal = st.slider("Dias no Matinal (últimos N dias do relatório de endereço)", 1, 7, 2)

    try:
        # Horas Extras
        raw = pd.read_excel(horas_file, sheet_name=0, header=None, dtype=str)
        long_df = detect_blocks_and_build_long(raw)
        df_he_final = clean_and_enrich(long_df)

        # Pontos com endereço
        df_addr_daily = read_pontos_endereco(endereco_file)

        st.success("✅ Arquivos processados com sucesso!")

        st.subheader("Prévia — Pontos com Endereço (diário)")
        st.dataframe(df_addr_daily.head(30), use_container_width=True)

        st.subheader("Prévia — HE (tratado)")
        st.dataframe(df_he_final[["Colaborador", "Data", "Semana", "ColabKey"]].head(30), use_container_width=True)

        excel_bytes = generate_excel_bytes(df_he_final, df_addr_daily, last_n_days_matinal)

        st.download_button(
            label="⬇️ Baixar Excel completo (HorasExtras_WS_Completo.xlsx)",
            data=excel_bytes,
            file_name="HorasExtras_WS_Completo.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"⚠️ Erro ao processar: {e}")


if __name__ == "__main__":
    main()
